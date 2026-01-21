"""
Temporal Stability Calculator for Video Pose Data

MODIFIED BEHAVIOR:
For each video and each joint:
- Evaluate all candidate frequencies
- Compute temporal jitter
- Select the frequency with least jitter
- Save selected frequency per (video, joint) into Excel
"""

import json
import os
import re
import numpy as np
import pandas as pd
import logging
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from enum import Enum

sys.path.insert(0, str(Path(__file__).parent.parent))
from utils.joint_enum import (
    GTJointsHumanSC3D,
    GTJointsHumanEVa,
    GTJointsMoVi,
    PredJointsDeepLabCut,
    Mediapipe,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


class TemporalStabilityCalculator:

    def __init__(self, video_json_folder: str, excel_file: str, joint_enum_class: Enum = None):

        self.video_json_folder = Path(video_json_folder)
        self.excel_file = Path(excel_file)
        self.joint_enum = joint_enum_class or PredJointsDeepLabCut

        if not self.video_json_folder.exists():
            raise ValueError(f"Video JSON folder does not exist: {video_json_folder}")

        if not self.excel_file.exists():
            raise ValueError(f"Excel file does not exist: {excel_file}")

        logger.info(f"Video JSON folder: {self.video_json_folder}")
        logger.info(f"Best frequency Excel: {self.excel_file}")
        logger.info(f"Joint enum: {self.joint_enum.__name__}")

    @staticmethod
    def get_joint_index_from_enum(joint_enum: Enum, joint_name: str) -> Optional[int]:
        try:
            joint_member = joint_enum[joint_name]
            index = joint_member.value
            if isinstance(index, tuple):
                return index[0]
            return int(index)
        except KeyError:
            return None

    def load_excel_frequencies(self) -> Dict[str, Dict[str, List[float]]]:
        """
        Returns:
        freq_dict[joint][video] = list of candidate freqs
        """

        xl = pd.ExcelFile(self.excel_file)

        target_sheet = None
        for sheet in xl.sheet_names:
            if "best" in sheet.lower() and "freq" in sheet.lower():
                target_sheet = sheet
                break
        if target_sheet is None:
            raise ValueError("No best-frequency sheet found")

        df = pd.read_excel(self.excel_file, sheet_name=target_sheet)

        freq_dict = {}  # joint -> video -> [freqs]

        for _, row in df.iterrows():
            video = str(row["Video"]).strip()

            for col in df.columns:
                joint = col.strip().upper()
                if joint == "VIDEO":
                    continue

                val = row[col]
                if pd.isna(val):
                    continue

                # parse frequencies in this cell
                freqs = []
                if isinstance(val, (int, float)):
                    freqs = [float(val)]
                elif isinstance(val, str):
                    parts = val.split(",")
                    for p in parts:
                        p = p.strip()
                        if p:
                            freqs.append(float(p))

                if not freqs:
                    continue

                freq_dict.setdefault(joint, {})
                freq_dict[joint][video] = sorted(set(freqs))

        logger.info("Loaded per-video candidate frequencies")
        return freq_dict


    @staticmethod
    def extract_frequency_from_folder(folder_path: Path) -> Optional[float]:
        match = re.search(r"(\d+)hz", folder_path.name, re.IGNORECASE)
        if match:
            return float(match.group(1))
        return None

    def find_video_json_files(self) -> Dict[float, Dict[str, Path]]:
        """
        Returns:
            freq_map[freq][video_name] = path_to_json_file
        """

        freq_map = {}

        # loop over filter_butterworth_*hz folders
        for freq_root in self.video_json_folder.iterdir():
            if not freq_root.is_dir():
                continue

            # parse frequency from folder name
            m = re.search(r'filter_butterworth_.*_(\d+)hz', freq_root.name, re.IGNORECASE)
            if not m:
                continue

            freq = float(m.group(1))

            # descend into timestamp folder then butterworth_order folder
            for date_folder in (freq_root / "filter").iterdir():
                if not date_folder.is_dir():
                    continue

                for bw_folder in date_folder.iterdir():
                    if not bw_folder.is_dir():
                        continue

                    # inside here: S1/S2/S3...
                    for subject_folder in bw_folder.iterdir():
                        if not subject_folder.is_dir():
                            continue

                        image_data = subject_folder / "Image_Data"
                        if not image_data.exists():
                            continue

                        # inside Image_Data: Box_1_(C2), Jog_1_(C1), ...
                        for action_cam_folder in image_data.iterdir():
                            if not action_cam_folder.is_dir():
                                continue

                            # find json file inside
                            json_files = list(action_cam_folder.glob("*.json"))
                            if not json_files:
                                continue

                            json_path = json_files[0]  # one per folder

                            # reconstruct video name to match Excel naming
                            subject = subject_folder.name  # S2
                            m2 = re.match(r'(.*)_\((C\d)\)', action_cam_folder.name)
                            if not m2:
                                continue

                            action = m2.group(1)      # Box_1
                            camera = m2.group(2)[1:]  # "2"

                            video_name = f"{subject}_{action}_{camera}.0"

                            freq_map.setdefault(freq, {})
                            freq_map[freq][video_name] = json_path

        logger.info(f"Discovered frequencies: {sorted(freq_map.keys())}")
        return freq_map


    @staticmethod
    def load_pose_data(json_path: Path) -> Dict:
        with open(json_path, "r") as f:
            return json.load(f)

    @staticmethod
    def extract_joint_xy_trajectory_normalized(pose_data: Dict, joint_index: int) -> np.ndarray:

        if "persons" not in pose_data or not pose_data["persons"]:
            return np.empty((0, 2))

        primary_person = max(pose_data["persons"], key=lambda p: len(p.get("poses", [])))
        poses = sorted(primary_person.get("poses", []), key=lambda p: p["frame_idx"])

        trajectory = []

        for pose in poses:
            kps = pose.get("keypoints", [])
            if isinstance(kps, list) and len(kps) == 1:
                kps = kps[0]

            if len(kps) <= joint_index:
                continue

            kp = kps[joint_index]
            if len(kp) < 2:
                continue

            x, y = float(kp[0]), float(kp[1])
            if not (np.isnan(x) or np.isnan(y)):
                trajectory.append([x, y])

        traj = np.asarray(trajectory, dtype=np.float32)
        if traj.shape[0] < 2:
            return traj

        min_xy = traj.min(axis=0)
        max_xy = traj.max(axis=0)
        scale = max_xy - min_xy
        scale[scale == 0] = 1.0

        return (traj - min_xy) / scale

    @staticmethod
    def compute_temporal_instability_euclidean(trajectory_xy: np.ndarray) -> float:

        if trajectory_xy.shape[0] < 2:
            return 0.0

        diffs = np.diff(trajectory_xy, axis=0)
        distances = np.linalg.norm(diffs, axis=1)
        return float(np.std(distances))

    def process_single_video_json(self, json_path: Path, joint_idx: int) -> float:

        pose_data = self.load_pose_data(json_path)
        traj_xy = self.extract_joint_xy_trajectory_normalized(pose_data, joint_idx)

        if traj_xy.shape[0] < 2:
            return 0.0

        return self.compute_temporal_instability_euclidean(traj_xy)

    def run_full_pipeline(self, output_excel: str):

        # now best_freq_dict[joint][video] = candidate list
        best_freq_dict = self.load_excel_frequencies()
        freq_to_files = self.find_video_json_files()

        best_results = {}   # joint -> video -> best freq
        all_jitters = {}    # freq -> joint -> video -> jitter

        for joint_name, video_freq_map in best_freq_dict.items():

            joint_idx = self.get_joint_index_from_enum(self.joint_enum, joint_name)
            if joint_idx is None:
                continue

            best_results[joint_name] = {}

            for video_name, candidate_freqs in video_freq_map.items():

                best_freq = None
                best_jitter = np.inf

                for freq in candidate_freqs:
                    if freq not in freq_to_files:
                        continue

                    # find correct json for this frequency + video
                    json_match = None
                    json_match = freq_to_files.get(freq, {}).get(video_name, None)
                    if json_match is None:
                        continue

                    jitter = self.process_single_video_json(json_match, joint_idx)

                    # store jitter
                    all_jitters.setdefault(freq, {})
                    all_jitters[freq].setdefault(joint_name, {})
                    all_jitters[freq][joint_name][video_name] = jitter

                    # pick best
                    if jitter < best_jitter:
                        best_jitter = jitter
                        best_freq = freq

                if best_freq is not None:
                    best_results[joint_name][video_name] = best_freq

        logger.info("Per-video per-joint frequency selection complete.")
        self.save_excel(all_jitters, best_results, output_excel)



    def save_excel(self, all_jitters: Dict, best_results: Dict, output_file: str):

        wb = Workbook()
        wb.remove(wb.active)

        # -----------------------------
        # 1) Create per-frequency jitter sheets
        # -----------------------------
        all_joints = sorted(best_results.keys())

        # collect all videos
        all_videos = set()
        for joint in best_results:
            all_videos.update(best_results[joint].keys())
        all_videos = sorted(all_videos)

        for freq in sorted(all_jitters.keys()):
            ws = wb.create_sheet(f"Freq_{int(freq)}Hz")

            ws.append(["Video"] + all_joints)

            for video in all_videos:
                row = [video]
                for joint in all_joints:
                    val = all_jitters.get(freq, {}).get(joint, {}).get(video, None)
                    if val is not None:
                        val = round(val, 8)
                    row.append(val)
                ws.append(row)

            self._apply_excel_formatting(ws)

        # -----------------------------
        # 2) Create best-frequency sheet
        # -----------------------------
        ws_best = wb.create_sheet("Best_Frequency", 0)
        ws_best.append(["Video"] + all_joints)

        for video in all_videos:
            row = [video]
            for joint in all_joints:
                val = best_results.get(joint, {}).get(video, None)
                row.append(val)
            ws_best.append(row)

        self._apply_excel_formatting(ws_best)

        wb.save(output_file)
        logger.info(f"Excel saved: {output_file}")



    @staticmethod
    def _apply_excel_formatting(worksheet):

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                val = str(cell.value)
                if val:
                    max_length = max(max_length, len(val))
            worksheet.column_dimensions[column_letter].width = max_length + 3


def main():

    video_json_folder = "/storageh100/Projects/Gaitly/bsehgal/pipeline_results/HumanEva/Butterworth_filter"
    best_freq_excel = "/storage/Projects/Gaitly/bsehgal/lower_body_pose_est/pipeline_results/Adaptive_filt/pck_summary.xlsx"
    output_excel = "/storage/Projects/Gaitly/bsehgal/lower_body_pose_est/pipeline_results/Adaptive_filt/temporal_stability_report.xlsx"

    joint_enum = PredJointsDeepLabCut

    calculator = TemporalStabilityCalculator(
        video_json_folder=video_json_folder,
        excel_file=best_freq_excel,
        joint_enum_class=joint_enum,
    )

    calculator.run_full_pipeline(output_excel)


if __name__ == "__main__":
    main()
